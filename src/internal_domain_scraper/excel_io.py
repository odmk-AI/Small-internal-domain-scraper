from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .config import SiteConfig
from .models import ScrapeResult


def normalize_header(value: Any) -> str:
    return str(value or "").strip().casefold()


def read_person_keys(input_xlsx: Path, sheet_name: str | None, config: SiteConfig) -> tuple[str, list[str]]:
    workbook = load_workbook(input_xlsx, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return config.id_headers[0], []

    accepted = {normalize_header(header) for header in config.id_headers}
    id_col = None
    id_header = config.id_headers[0]
    for index, header in enumerate(rows[0]):
        if normalize_header(header) in accepted:
            id_col = index
            id_header = str(header or config.id_headers[0]).strip() or config.id_headers[0]
            break

    if id_col is None:
        accepted_list = ", ".join(config.id_headers)
        raise ValueError(f"Keine ID-Spalte gefunden. Erlaubte Kopfzeilen: {accepted_list}")

    keys: list[str] = []
    seen: set[str] = set()
    for row in rows[1:]:
        raw = row[id_col] if id_col < len(row) else None
        person_key = str(raw or "").strip()
        if not re.fullmatch(r"\d+", person_key):
            continue
        if person_key not in seen:
            seen.add(person_key)
            keys.append(person_key)

    return id_header, keys


def write_output_xlsx(output_xlsx: Path, id_header: str, config: SiteConfig, results: list[ScrapeResult]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scraped Values"
    headers = [id_header, *[field.output_header for field in config.fields]]
    sheet.append(headers)

    for result in results:
        sheet.append([result.person_key, *[result.values.get(field.key) for field in config.fields]])

    for cell in sheet["A"]:
        cell.number_format = "@"

    for column_index, field in enumerate(config.fields, start=2):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        if field.type == "percent":
            for cell in sheet[column_letter][1:]:
                cell.number_format = "0%"
        sheet.column_dimensions[column_letter].width = max(20, len(field.output_header) + 2)

    sheet.column_dimensions["A"].width = max(16, len(id_header) + 2)
    if config.mode == "taskboard":
        _append_taskboard_hours_summary(workbook, results)
    workbook.save(output_xlsx)


def write_output_csv(output_csv: Path, id_header: str, config: SiteConfig, results: list[ScrapeResult]) -> None:
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow([id_header, *[field.output_header for field in config.fields]])
        for result in results:
            row: list[str] = [result.person_key]
            for field in config.fields:
                value = result.values.get(field.key)
                if value is None:
                    row.append("")
                elif field.type == "percent" and isinstance(value, (float, int)):
                    row.append(f"{value:.6f}")
                else:
                    row.append(str(value))
            writer.writerow(row)


def _append_taskboard_hours_summary(workbook: Workbook, results: list[ScrapeResult]) -> None:
    sheet = workbook.create_sheet("Hours per Sprint")
    headers = ["Assigned To", "Year", "Sprint", "Sum Completed", "Sum Open"]
    sheet.append(headers)

    totals: dict[tuple[str, str, str], dict[str, float]] = {}
    for result in results:
        assigned_to = str(result.values.get("assigned_to") or "").strip()
        sprint = str(result.values.get("sprint") or "").strip()
        year = str(result.values.get("year") or "").strip()
        if not assigned_to and not sprint:
            continue

        key = (assigned_to, year, sprint)
        if key not in totals:
            totals[key] = {"completed": 0.0, "open": 0.0}
        totals[key]["completed"] += _parse_hours(result.values.get("completed_work"))
        totals[key]["open"] += _parse_hours(result.values.get("remaining_work"))

    for assigned_to, year, sprint in sorted(totals):
        sheet.append(
            [
                assigned_to,
                year,
                sprint,
                totals[(assigned_to, year, sprint)]["completed"],
                totals[(assigned_to, year, sprint)]["open"],
            ]
        )

    for column_index, header in enumerate(headers, start=1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = max(14, len(header) + 2)

    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "0.00"


def _parse_hours(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (float, int)):
        return float(value)

    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    return float(match.group(0))
